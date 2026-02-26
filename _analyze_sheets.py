import openpyxl

wb = openpyxl.load_workbook("ReporteGasto_30.01.26.xlsx", data_only=True)
print("=" * 120)
print("SHEETS IN WORKBOOK:", wb.sheetnames)
print("=" * 120)

sf = wb["SF"]
print("\n" + "=" * 120)
print("SF SHEET ANALYSIS")
print("=" * 120)
print("Total rows:", sf.max_row, "Total cols:", sf.max_column)

print("\n--- ROW 1 (Header line 1) ---")
for c in range(1, sf.max_column + 1):
    print("  Col %2d: %s" % (c, sf.cell(row=1, column=c).value))

print("\n--- ROW 2 (Header line 2) ---")
for c in range(1, sf.max_column + 1):
    print("  Col %2d: %s" % (c, sf.cell(row=2, column=c).value))

expected = {1:"sec_func",2:"nro_pp",3:"programa_pptal",4:"tipo",5:"producto_proyecto",6:"actividad",7:"funcion",8:"division_func",9:"grupo_func",10:"finalidad",11:"organo",12:"unidad_org",13:"sub_unidad",14:"nombre_corto"}

print("\n--- COLUMN MAPPING VERIFICATION ---")
for ci, fn in expected.items():
    h1 = sf.cell(row=1, column=ci).value
    h2 = sf.cell(row=2, column=ci).value
    print("  Col %2d | %-20s | H1: %-50s | H2: %-50s" % (ci, fn, str(h1)[:50], str(h2)[:50]))

print("\n--- FIRST DATA ROW (Row 3) - Full detail ---")
for c in range(1, sf.max_column + 1):
    v = sf.cell(row=3, column=c).value
    h1 = sf.cell(row=1, column=c).value
    h2 = sf.cell(row=2, column=c).value
    print("  Col %2d (%s / %s): %s" % (c, h1, h2, v))

valid_rows = []
invalid_rows = []
empty_rows = []

for ri in range(3, sf.max_row + 1):
    rd = [sf.cell(row=ri, column=c).value for c in range(1, sf.max_column + 1)]
    if all(v is None for v in rd):
        empty_rows.append(ri)
        continue
    bad = False
    for v in rd:
        if v is not None and isinstance(v, str) and ("#REF!" in v or "ANULADO" in v.upper()):
            bad = True
            break
    if bad:
        invalid_rows.append((ri, rd))
    else:
        valid_rows.append((ri, rd))

print("\nTotal non-empty rows:", len(valid_rows) + len(invalid_rows))
print("Valid rows:", len(valid_rows))
print("Invalid rows (#REF!/ANULADO):", len(invalid_rows))
print("Empty rows:", len(empty_rows))

print("\n--- ALL VALID SF ROWS ---")
print("  %4s | %-12s | %-8s | %-35s | %-6s | %-35s | %-45s" % ("Row","sec_func","nro_pp","programa_pptal","tipo","producto_proyecto","nombre_corto"))
print("  " + "-" * 155)
for ri, rd in valid_rows:
    sf0 = str(rd[0])[:12] if rd[0] is not None else ""
    sf1 = str(rd[1])[:8] if rd[1] is not None else ""
    sf2 = str(rd[2])[:35] if rd[2] is not None else ""
    sf3 = str(rd[3])[:6] if rd[3] is not None else ""
    sf4 = str(rd[4])[:35] if rd[4] is not None else ""
    sf13 = str(rd[13])[:45] if len(rd) > 13 and rd[13] is not None else ""
    print("  %4d | %-12s | %-8s | %-35s | %-6s | %-35s | %-45s" % (ri, sf0, sf1, sf2, sf3, sf4, sf13))

if invalid_rows:
    print("\n--- INVALID SF ROWS ---")
    for ri, rd in invalid_rows:
        print("  Row %d: %s" % (ri, [str(v)[:30] if v is not None else "" for v in rd[:14]]))

sf_metas = set()
for ri, rd in valid_rows:
    if rd[0] is not None:
        sf_metas.add(str(rd[0]).strip())

print("\n--- UNIQUE sec_func (METAS) IN SF: %d ---" % len(sf_metas))
for m in sorted(sf_metas):
    print("  " + m)

if sf.max_column > 14:
    print("\n--- ADDITIONAL COLUMNS BEYOND 14 ---")
    for c in range(15, sf.max_column + 1):
        h1 = sf.cell(row=1, column=c).value
        h2 = sf.cell(row=2, column=c).value
        sample = sf.cell(row=3, column=c).value
        print("  Col %d: H1=%s, H2=%s, Sample=%s" % (c, h1, h2, sample))

# SheetGasto
print("\n" + "=" * 120)
print("SheetGasto ANALYSIS")
print("=" * 120)
sg = wb["SheetGasto"]
print("Total rows:", sg.max_row, "Total cols:", sg.max_column)
print("\n--- SheetGasto Headers (Row 1) ---")
for c in range(1, min(sg.max_column + 1, 25)):
    print("  Col %2d: %s" % (c, sg.cell(row=1, column=c).value))

print("\n--- Scanning for meta-like columns ---")
sg_metas = set()
for c in range(1, min(sg.max_column + 1, 25)):
    header = sg.cell(row=1, column=c).value
    vals = set()
    for ri in range(2, min(sg.max_row + 1, 500)):
        v = sg.cell(row=ri, column=c).value
        if v is not None:
            vals.add(str(v).strip())
    if 0 < len(vals) <= 30:
        print("    Col %d (%s): %d unique -> %s" % (c, header, len(vals), sorted(vals)[:20]))
    if header and "sec" in str(header).lower():
        sg_metas = vals

if not sg_metas:
    for ri in range(2, sg.max_row + 1):
        v = sg.cell(row=ri, column=1).value
        if v is not None:
            sg_metas.add(str(v).strip())

# RB
print("\n" + "=" * 120)
print("RB SHEET ANALYSIS")
print("=" * 120)
rb = wb["RB"]
print("Total rows:", rb.max_row, "Total cols:", rb.max_column)
print("\n--- RB Headers (Row 1) ---")
for c in range(1, rb.max_column + 1):
    print("  Col %2d: %s" % (c, rb.cell(row=1, column=c).value))
print("\n--- RB Headers (Row 2) ---")
for c in range(1, rb.max_column + 1):
    print("  Col %2d: %s" % (c, rb.cell(row=2, column=c).value))
print("\n--- ALL RB DATA ROWS ---")
rb_v = 0
rb_i = 0
rb_e = 0
for ri in range(3, rb.max_row + 1):
    rv = [rb.cell(row=ri, column=c).value for c in range(1, rb.max_column + 1)]
    if all(v is None for v in rv):
        rb_e += 1
        continue
    bad = any(v is not None and isinstance(v, str) and ("#REF!" in v or "ANULADO" in str(v).upper()) for v in rv)
    if bad:
        rb_i += 1
        tag = " [INVALID]"
    else:
        rb_v += 1
        tag = ""
    disp = [str(v)[:30] if v is not None else "---" for v in rv[:12]]
    extra = " +%d more" % (len(rv)-12) if len(rv) > 12 else ""
    print("  Row %3d%s: %s%s" % (ri, tag, disp, extra))
print("\n  RB Summary: valid=%d, invalid=%d, empty=%d" % (rb_v, rb_i, rb_e))

# CG
print("\n" + "=" * 120)
print("CG SHEET ANALYSIS")
print("=" * 120)
cg = wb["CG"]
print("Total rows:", cg.max_row, "Total cols:", cg.max_column)
print("\n--- CG Headers (Row 1) ---")
for c in range(1, cg.max_column + 1):
    print("  Col %2d: %s" % (c, cg.cell(row=1, column=c).value))
print("\n--- CG Headers (Row 2) ---")
for c in range(1, cg.max_column + 1):
    print("  Col %2d: %s" % (c, cg.cell(row=2, column=c).value))

cg_data = []
for ri in range(3, cg.max_row + 1):
    rv = [cg.cell(row=ri, column=c).value for c in range(1, cg.max_column + 1)]
    if not all(v is None for v in rv):
        cg_data.append((ri, rv))

print("\nTotal non-empty CG data rows:", len(cg_data))
print("\n--- CG SAMPLE ROWS (first 8 cols) ---")
if len(cg_data) <= 50:
    samples = cg_data
else:
    mid = len(cg_data) // 2
    samples = cg_data[:20] + [(-1, ["... MIDDLE ..."])] + cg_data[mid-2:mid+3] + [(-1, ["... END ..."])] + cg_data[-10:]

for ri, rv in samples:
    if ri == -1:
        print("  " + rv[0])
        continue
    disp = [str(v)[:35] if v is not None else "---" for v in rv[:8]]
    print("  Row %4d: %s" % (ri, disp))

# COMPARISON
print("\n" + "=" * 120)
print("FINAL COMPARISON: SF vs SheetGasto METAS")
print("=" * 120)
print("  SF valid metas:", len(sf_metas))
print("  SheetGasto metas:", len(sg_metas))
print("  SF metas sorted:", sorted(sf_metas))
print("  SheetGasto metas sorted:", sorted(sg_metas))
if sg_metas:
    both = sf_metas & sg_metas
    sf_only = sf_metas - sg_metas
    sg_only = sg_metas - sf_metas
    print("  In BOTH (%d): %s" % (len(both), sorted(both)))
    print("  SF ONLY (%d): %s" % (len(sf_only), sorted(sf_only)))
    print("  SheetGasto ONLY (%d): %s" % (len(sg_only), sorted(sg_only)))

wb.close()
print("\nDone.")
