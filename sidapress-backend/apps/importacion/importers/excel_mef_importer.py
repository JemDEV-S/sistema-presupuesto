import logging
from decimal import Decimal, InvalidOperation
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from apps.catalogos.models import AnioFiscal, FuenteFinanciamiento, Rubro, ClasificadorGasto
from apps.organizacion.models import UnidadOrganica
from apps.presupuesto.models import Meta, EjecucionPresupuestal, EjecucionMensual, AvanceFisico
from apps.importacion.models import ImportacionArchivo
from apps.alertas.models import Alerta
from apps.alertas.generators.alert_generator import generar_alertas

logger = logging.getLogger(__name__)

# =============================================================================
# Mapeo de columnas para .xlsx (107 columnas, con columnas intermedias)
# =============================================================================
COL_XLSX = {
    'ano_eje': 0,
    'programa_pptal': 5,
    'tipo_prod_proy': 6,
    'producto_proyecto': 7,
    'tipo_act_obra_ac': 8,
    'activ_obra_accinv': 9,
    'funcion': 10,
    'division_fn': 11,
    'grupo_fn': 12,
    'meta': 13,
    'finalidad': 14,
    'unidad_medida': 15,
    'cant_meta_anual': 16,
    'cant_meta_sem': 17,
    'avan_fisico_anual': 18,
    'avan_fisico_sem': 19,
    'sec_func': 20,
    'fuente_financ': 24,
    'rubro': 25,
    'categoria_gasto': 26,
    'tipo_transaccion': 27,
    # En .xlsx las columnas de clasificador están separadas en codigo y nombre
    'generica_cod': 29,
    'generica_nom': 30,
    'subgenerica_cod': 31,
    'subgenerica_nom': 32,
    'subgenerica_det_cod': 33,
    'subgenerica_det_nom': 34,
    'especifica_cod': 35,
    'especifica_nom': 36,
    'especifica_det_cod': 37,
    'especifica_det_nom': 38,
    'mto_pia': 39,
    'mto_modificaciones': 40,
    'mto_pim': 41,
    'mto_certificado': 42,
    'mto_compro_anual': 43,
    'mto_at_comp_01': 44,
    'mto_devenga_01': 56,
    'mto_girado_01': 68,
    'mto_pagado_01': 80,
    'filtro_restringido': 101,
}

# =============================================================================
# Mapeo de columnas para .xls diario (86 columnas, formato compacto)
# =============================================================================
COL_DAILY = {
    'ano_eje': 0,
    'programa_pptal': 5,
    'tipo_prod_proy': 6,
    'producto_proyecto': 7,
    'tipo_act_obra_ac': 8,
    'activ_obra_accinv': 9,
    'funcion': 10,
    'division_fn': 11,
    'grupo_fn': 12,
    'meta': 13,
    'finalidad': 14,
    'unidad_medida': 15,
    'cant_meta_anual': 16,
    'cant_meta_sem': 17,
    'avan_fisico_anual': 18,
    'avan_fisico_sem': 19,
    'sec_func': 20,
    'fuente_financ': 24,
    'rubro': 25,
    'categoria_gasto': 26,
    'tipo_transaccion': 27,
    'generica': 28,
    'subgenerica': 29,
    'subgenerica_det': 30,
    'especifica': 31,
    'especifica_det': 32,
    'mto_pia': 33,
    'mto_modificaciones': 34,
    'mto_pim': 35,
    'mto_certificado': 36,
    'mto_compro_anual': 37,
    'mto_at_comp_01': 38,
    'mto_devenga_01': 50,
    'mto_girado_01': 62,
    'mto_pagado_01': 74,
}

# Columnas de la hoja SF (0-indexed, hoja empieza en fila 3)
COL_SF = {
    'sec_func': 1,
    'nro_pp': 2,
    'programa_pptal': 3,
    'tipo': 4,
    'producto_proyecto': 5,
    'actividad': 6,
    'funcion': 7,
    'division_func': 8,
    'grupo_func': 9,
    'finalidad': 10,
    'organo': 11,
    'unidad_org': 12,
    'sub_unidad': 13,
    'nombre_corto': 14,
}

# Columnas de la hoja RB
COL_RB = {
    'anio': 0,
    'fuente': 1,
    'rubro': 2,
    'des_rubro': 3,
    'nombre_corto': 4,
}

# Columnas de la hoja CG
COL_CG = {
    'anio': 0,
    'clasificador': 1,
    'descripcion': 2,
    'descripcion_detallada': 3,
}


# =============================================================================
# Helpers
# =============================================================================

def parse_codigo_nombre(value):
    """
    Extrae código y nombre de formato "CODIGO.DESCRIPCION".
    "0002.SALUD MATERNO NEONATAL" → ("0002", "SALUD MATERNO NEONATAL")
    "" o None → ("", "")
    """
    if not value:
        return ('', '')
    s = str(value).strip()
    if '.' in s:
        parts = s.split('.', 1)
        return (parts[0].strip(), parts[1].strip())
    return (s, '')


def safe_decimal(value):
    if value is None:
        return Decimal('0')
    try:
        return Decimal(str(value)).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        return Decimal('0')


def safe_str(value):
    if value is None:
        return ''
    return str(value).strip()


def safe_int(value):
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def get_cell(row, col_index):
    if col_index < len(row):
        return row[col_index]
    return None


def build_ejecucion_key(meta_codigo, rubro_codigo, clasificador_codigo):
    return f"{meta_codigo}|{rubro_codigo}|{clasificador_codigo}"


def montos_changed(existing_obj, new_data):
    """Compara si algún monto cambió entre el registro existente y los datos nuevos."""
    for field in ('pia', 'modificaciones', 'pim', 'certificado', 'compromiso_anual'):
        if getattr(existing_obj, field) != new_data[field]:
            return True
    return False


def mensual_changed(existing_obj, new_data):
    """Compara si algún campo mensual cambió."""
    for field in ('compromiso', 'devengado', 'girado', 'pagado'):
        if getattr(existing_obj, field) != new_data[field]:
            return True
    return False


def _is_invalid_sf_value(value):
    """Detecta valores inválidos en la hoja SF (#REF!, ANULADO, vacíos)."""
    if not value:
        return True
    s = str(value).strip().upper()
    return '#REF' in s or s == 'ANULADO' or s == ''


def _normalize_name(value):
    """Normaliza un nombre para evitar duplicados por espacios o capitalización."""
    if not value:
        return ''
    return ' '.join(str(value).strip().upper().split())


def _normalize_tipo_actividad(raw):
    """Mapea valor Excel tipo_act_obra_ac a choice del modelo Meta.TIPO_ACTIVIDAD."""
    if not raw:
        return ''
    val = raw.strip().upper()
    if 'OBRA' in val:
        return 'OBRA'
    if 'ACCION' in val or 'ACCIÓN' in val or 'INVERSIÓN' in val or 'INVERSION' in val:
        return 'ACCION_INVERSION'
    return 'ACTIVIDAD'


# =============================================================================
# Procesadores de hojas (carga inicial .xlsx)
# =============================================================================

def _process_rb_sheet(wb):
    """
    RB → FuenteFinanciamiento + Rubro (padre-hijo).
    Retorna dict {rubro_codigo: rubro_obj}.
    """
    if 'RB' not in wb.sheetnames:
        logger.warning('Hoja RB no encontrada, saltando')
        return {}

    ws = wb['RB']
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    rubros_cache = {}
    fuentes_cache = {}  # fuente_cod → FuenteFinanciamiento obj

    for row in rows:
        try:
            fuente_raw = safe_str(get_cell(row, COL_RB['fuente']))
            rubro_raw = safe_str(get_cell(row, COL_RB['rubro']))
            des_rubro = safe_str(get_cell(row, COL_RB['des_rubro']))
            nombre_corto_raw = safe_str(get_cell(row, COL_RB['nombre_corto']))

            fuente_cod, fuente_nom = parse_codigo_nombre(fuente_raw)
            rubro_cod, rubro_nom = parse_codigo_nombre(rubro_raw)

            if not fuente_cod or not rubro_cod:
                continue

            # Crear/obtener FuenteFinanciamiento (dedup por fuente_cod)
            if fuente_cod not in fuentes_cache:
                fuente_obj, _ = FuenteFinanciamiento.objects.get_or_create(
                    codigo=fuente_cod[:10],
                    defaults={
                        'nombre': fuente_nom[:200],
                        'nombre_corto': fuente_cod,
                    }
                )
                fuentes_cache[fuente_cod] = fuente_obj

            fuente_obj = fuentes_cache[fuente_cod]

            # Crear/obtener Rubro (hijo de la fuente)
            rubro_obj, created = Rubro.objects.update_or_create(
                codigo=rubro_cod[:10],
                defaults={
                    'nombre': rubro_nom[:200],
                    'nombre_corto': nombre_corto_raw[:50],
                    'descripcion': des_rubro,
                    'fuente': fuente_obj,
                }
            )
            rubros_cache[rubro_cod] = rubro_obj

        except Exception as e:
            logger.warning(f'Error procesando fila RB: {e}')

    logger.info(f'RB: {len(fuentes_cache)} fuentes, {len(rubros_cache)} rubros procesados')
    return rubros_cache


def _process_cg_sheet(wb):
    """
    CG → ClasificadorGasto (todos los niveles).
    Parsea las 6 partes del código incluyendo subgenerica_det y especifica_det.
    Retorna dict {codigo: obj}.
    """
    if 'CG' not in wb.sheetnames:
        logger.warning('Hoja CG no encontrada, saltando')
        return {}

    ws = wb['CG']
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    existing_codigos = set(ClasificadorGasto.objects.values_list('codigo', flat=True))
    to_create = []

    for row in rows:
        try:
            codigo = safe_str(get_cell(row, COL_CG['clasificador']))
            descripcion = safe_str(get_cell(row, COL_CG['descripcion']))
            descripcion_det = safe_str(get_cell(row, COL_CG['descripcion_detallada']))

            if not codigo or codigo in existing_codigos:
                continue

            parts = codigo.split('.')
            to_create.append(ClasificadorGasto(
                codigo=codigo[:20],
                nombre=descripcion[:300],
                tipo_transaccion=parts[0] if len(parts) > 0 else '',
                generica=parts[1] if len(parts) > 1 else '',
                subgenerica=parts[2] if len(parts) > 2 else '',
                subgenerica_det=parts[3] if len(parts) > 3 else '',
                especifica=parts[4] if len(parts) > 4 else '',
                especifica_det=parts[5] if len(parts) > 5 else '',
                descripcion_detallada=descripcion_det,
            ))
            existing_codigos.add(codigo)
        except Exception as e:
            logger.warning(f'Error procesando fila CG: {e}')

    if to_create:
        ClasificadorGasto.objects.bulk_create(to_create, batch_size=200)

    cache = {c.codigo: c for c in ClasificadorGasto.objects.all()}
    logger.info(f'CG: {len(to_create)} clasificadores creados, {len(cache)} total')
    return cache


def _enrich_clasificadores_from_sheetgasto(wb, clasificadores_cache):
    """
    Cruza datos de SheetGasto para llenar los nombres por nivel de los clasificadores.
    SheetGasto tiene columnas separadas de código/nombre para cada nivel.
    Solo actualiza clasificadores que existen en SheetGasto.
    """
    if 'SheetGasto' not in wb.sheetnames:
        return

    ws = wb['SheetGasto']
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Recolectar nombres por clasificador_codigo
    nombres_por_codigo = {}
    for row in rows:
        try:
            tipo_trans_raw = safe_str(get_cell(row, COL_XLSX['tipo_transaccion']))
            tipo_trans_cod, tipo_trans_nom = parse_codigo_nombre(tipo_trans_raw)
            gen_cod = safe_str(get_cell(row, COL_XLSX['generica_cod']))
            gen_nom = safe_str(get_cell(row, COL_XLSX['generica_nom']))
            subgen_cod = safe_str(get_cell(row, COL_XLSX['subgenerica_cod']))
            subgen_nom = safe_str(get_cell(row, COL_XLSX['subgenerica_nom']))
            subgen_det_cod = safe_str(get_cell(row, COL_XLSX['subgenerica_det_cod']))
            subgen_det_nom = safe_str(get_cell(row, COL_XLSX['subgenerica_det_nom']))
            esp_cod = safe_str(get_cell(row, COL_XLSX['especifica_cod']))
            esp_nom = safe_str(get_cell(row, COL_XLSX['especifica_nom']))
            esp_det_cod = safe_str(get_cell(row, COL_XLSX['especifica_det_cod']))
            esp_det_nom = safe_str(get_cell(row, COL_XLSX['especifica_det_nom']))

            codigo = f'{tipo_trans_cod}.{gen_cod}.{subgen_cod}.{subgen_det_cod}.{esp_cod}.{esp_det_cod}'

            if codigo not in nombres_por_codigo:
                nombres_por_codigo[codigo] = {
                    'nombre_tipo_transaccion': tipo_trans_nom[:100],
                    'nombre_generica': gen_nom[:200],
                    'nombre_subgenerica': subgen_nom[:200],
                    'nombre_subgenerica_det': subgen_det_nom[:200],
                    'nombre_especifica': esp_nom[:200],
                    'nombre_especifica_det': esp_det_nom[:300],
                }
        except Exception:
            continue

    # Actualizar clasificadores en BD
    to_update = []
    for codigo, nombres in nombres_por_codigo.items():
        if codigo in clasificadores_cache:
            obj = clasificadores_cache[codigo]
            changed = False
            for field, value in nombres.items():
                if value and getattr(obj, field) != value:
                    setattr(obj, field, value)
                    changed = True
            if changed:
                to_update.append(obj)

    if to_update:
        ClasificadorGasto.objects.bulk_update(
            to_update,
            ['nombre_tipo_transaccion', 'nombre_generica', 'nombre_subgenerica',
             'nombre_subgenerica_det', 'nombre_especifica', 'nombre_especifica_det'],
            batch_size=200,
        )

    logger.info(f'Clasificadores enriquecidos con nombres de SheetGasto: {len(to_update)} actualizados')


def _make_unit_codigo(nombre, prefix='', hierarchy_parts=None):
    """Genera un código único para UnidadOrganica basado en la jerarquía completa."""
    import hashlib
    if hierarchy_parts is None:
        hierarchy_parts = []
    clean = _normalize_name(nombre)
    key = '|'.join([_normalize_name(p) for p in hierarchy_parts] + [clean])
    h = hashlib.md5(key.encode()).hexdigest()[:12]
    return f'{prefix}{h}'[:20]


def _process_sf_sheet(wb, anio_fiscal):
    """
    SF → UnidadOrganica (jerarquía) + Meta (una por sec_func).
    Retorna dict {sec_func: {'unidad': unidad_obj, 'meta': meta_obj}}.
    Filtra filas con #REF! y ANULADO. Normaliza nombres para evitar duplicados.
    """
    if 'SF' not in wb.sheetnames:
        logger.warning('Hoja SF no encontrada, saltando')
        return {}

    ws = wb['SF']
    rows = list(ws.iter_rows(min_row=3, values_only=True))

    sf_mapping = {}
    organos_cache = {}  # nombre_normalizado → obj
    unidades_cache = {}  # (nombre_organo_norm, nombre_unidad_norm) → obj
    subunidades_cache = {}  # (nombre_organo_norm, nombre_unidad_norm, nombre_sub_norm) → obj
    skipped_invalid = 0

    for row_num, row in enumerate(rows, start=3):
        try:
            sec_func = safe_int(get_cell(row, COL_SF['sec_func']))
            if sec_func is None:
                continue

            organo_nombre = safe_str(get_cell(row, COL_SF['organo']))
            unidad_nombre = safe_str(get_cell(row, COL_SF['unidad_org']))
            sub_unidad_nombre = safe_str(get_cell(row, COL_SF['sub_unidad']))
            nombre_corto = safe_str(get_cell(row, COL_SF['nombre_corto']))

            # Filtrar filas ANULADO y #REF! (no se pueden usar)
            organo_invalid = _is_invalid_sf_value(organo_nombre)
            if organo_nombre and not organo_invalid:
                pass  # organo válido
            elif not organo_nombre or organo_nombre is None:
                pass  # organo vacío - meta sin unidad orgánica
            else:
                # #REF! o ANULADO
                skipped_invalid += 1
                continue

            # Crear jerarquía de UnidadOrganica solo si hay organo válido
            organo_obj = None
            unidad_obj = None
            sub_unidad_obj = None
            unidad_asignada = None

            if not organo_invalid and organo_nombre:
                organo_norm = _normalize_name(organo_nombre)
                unidad_norm = _normalize_name(unidad_nombre)
                sub_unidad_norm = _normalize_name(sub_unidad_nombre)

                # Nivel 1: Órgano
                if organo_norm not in organos_cache:
                    org_codigo = _make_unit_codigo(organo_nombre, 'O_')
                    organo_obj, _ = UnidadOrganica.objects.get_or_create(
                        nombre=organo_nombre[:200],
                        nivel=1,
                        parent=None,
                        defaults={
                            'codigo': org_codigo,
                            'nombre_corto': organo_nombre[:50],
                        }
                    )
                    organos_cache[organo_norm] = organo_obj
                organo_obj = organos_cache[organo_norm]

                # Nivel 2: Unidad Orgánica
                if unidad_nombre and not _is_invalid_sf_value(unidad_nombre):
                    uo_key = (organo_norm, unidad_norm)
                    if uo_key not in unidades_cache:
                        uo_codigo = _make_unit_codigo(unidad_nombre, 'U_', [organo_nombre])
                        unidad_obj, _ = UnidadOrganica.objects.get_or_create(
                            nombre=unidad_nombre[:200],
                            nivel=2,
                            parent=organo_obj,
                            defaults={
                                'codigo': uo_codigo,
                                'nombre_corto': nombre_corto[:50] if nombre_corto else unidad_nombre[:50],
                            }
                        )
                        unidades_cache[uo_key] = unidad_obj
                    unidad_obj = unidades_cache[uo_key]

                # Nivel 3: Sub Unidad Orgánica
                if sub_unidad_nombre and unidad_obj and not _is_invalid_sf_value(sub_unidad_nombre):
                    sub_key = (organo_norm, unidad_norm, sub_unidad_norm)
                    if sub_key not in subunidades_cache:
                        sub_codigo = _make_unit_codigo(sub_unidad_nombre, 'S_', [organo_nombre, unidad_nombre])
                        sub_unidad_obj, _ = UnidadOrganica.objects.get_or_create(
                            nombre=sub_unidad_nombre[:200],
                            nivel=3,
                            parent=unidad_obj,
                            defaults={
                                'codigo': sub_codigo,
                                'nombre_corto': nombre_corto[:50] if nombre_corto else sub_unidad_nombre[:50],
                            }
                        )
                        subunidades_cache[sub_key] = sub_unidad_obj
                    sub_unidad_obj = subunidades_cache[sub_key]

                # Asignar al nivel más bajo disponible
                unidad_asignada = sub_unidad_obj or unidad_obj or organo_obj

            # ---- Crear Meta desde SF ----
            meta_codigo = str(sec_func).zfill(4)

            # Parsear campos con formato "CODIGO.NOMBRE"
            nro_pp = safe_str(get_cell(row, COL_SF['nro_pp']))
            programa_pptal_raw = safe_str(get_cell(row, COL_SF['programa_pptal']))
            tipo_raw = safe_str(get_cell(row, COL_SF['tipo']))
            producto_proyecto_raw = safe_str(get_cell(row, COL_SF['producto_proyecto']))
            actividad_raw = safe_str(get_cell(row, COL_SF['actividad']))
            funcion_raw = safe_str(get_cell(row, COL_SF['funcion']))
            division_func_raw = safe_str(get_cell(row, COL_SF['division_func']))
            grupo_func_raw = safe_str(get_cell(row, COL_SF['grupo_func']))
            finalidad_raw = safe_str(get_cell(row, COL_SF['finalidad']))

            # Parsear código.nombre de los campos que lo tienen
            prod_cod, prod_nom = parse_codigo_nombre(producto_proyecto_raw)
            act_cod, act_nom = parse_codigo_nombre(actividad_raw)
            func_cod, func_nom = parse_codigo_nombre(funcion_raw)
            div_cod, div_nom = parse_codigo_nombre(division_func_raw)
            grupo_cod, grupo_nom = parse_codigo_nombre(grupo_func_raw)
            final_cod, final_nom = parse_codigo_nombre(finalidad_raw)

            # Determinar nombre de la meta
            nombre_meta = act_nom or prod_nom or programa_pptal_raw or f'Meta {meta_codigo}'

            # Determinar tipo (tipo_raw viene de la columna tipo_prod_proy: PRODUCTO o PROYECTO)
            tipo_meta = 'PROYECTO' if tipo_raw and tipo_raw.upper() == 'PROYECTO' else 'PRODUCTO'
            tipo_prod_proy = tipo_raw if tipo_raw else ''

            meta_obj, _ = Meta.objects.update_or_create(
                anio_fiscal=anio_fiscal,
                codigo=meta_codigo,
                defaults={
                    'nombre': nombre_meta[:500],
                    'unidad_organica': unidad_asignada,
                    'tipo_meta': tipo_meta,
                    'cantidad_meta_anual': Decimal('0'),
                    'finalidad': finalidad_raw[:500] if finalidad_raw else '',
                    # Códigos numéricos
                    'sec_func': sec_func,
                    'codigo_programa_pptal': nro_pp[:20] if nro_pp else '',
                    'codigo_producto_proyecto': prod_cod[:20],
                    'codigo_actividad': act_cod[:20],
                    'codigo_funcion': func_cod[:10],
                    'codigo_division_fn': div_cod[:10],
                    'codigo_grupo_fn': grupo_cod[:10],
                    'codigo_finalidad': final_cod[:20],
                    # Nombres
                    'nombre_programa_pptal': programa_pptal_raw[:300] if programa_pptal_raw else '',
                    'nombre_producto_proyecto': prod_nom[:500],
                    'nombre_actividad': act_nom[:500],
                    # Tipos y clasificación
                    'tipo_producto_proyecto': tipo_prod_proy[:50],
                    'tipo_actividad': '',  # Se actualizará desde SheetGasto con tipo_act_obra_ac
                    'codigo_unidad_medida': '',
                    'nombre_unidad_medida': '',
                }
            )

            # Crear AvanceFisico con valores iniciales 0 (se actualizará desde SheetGasto)
            AvanceFisico.objects.get_or_create(
                meta=meta_obj,
                defaults={
                    'cantidad_meta_semestral': Decimal('0'),
                    'avance_fisico_anual': Decimal('0'),
                    'avance_fisico_semestral': Decimal('0'),
                }
            )

            sf_mapping[sec_func] = {
                'unidad': unidad_asignada,
                'meta': meta_obj,
            }

        except Exception as e:
            logger.warning(f'Error procesando SF fila {row_num}: {e}')

    logger.info(f'SF: {len(sf_mapping)} sec_func mapeadas, '
                f'{len(organos_cache)} órganos, {len(unidades_cache)} unidades, '
                f'{len(subunidades_cache)} sub-unidades, '
                f'{skipped_invalid} filas inválidas filtradas')
    return sf_mapping


# =============================================================================
# Procesamiento de SheetGasto (compartido por ambos flujos)
# =============================================================================

def _extract_row_data(row, col_map, is_xlsx=False):
    """
    Extrae y parsea todos los datos de una fila de SheetGasto.
    Maneja tanto formato .xlsx (107 cols) como .xls (86 cols).
    """
    data = {}

    # Campos simples
    data['meta_num'] = safe_int(get_cell(row, col_map['meta']))
    data['sec_func'] = safe_int(get_cell(row, col_map['sec_func']))
    data['cant_meta_anual'] = safe_decimal(get_cell(row, col_map['cant_meta_anual']))

    # Avance físico (cols 17-19, ambos formatos)
    data['cant_meta_sem'] = safe_decimal(get_cell(row, col_map['cant_meta_sem']))
    data['avan_fisico_anual'] = safe_decimal(get_cell(row, col_map['avan_fisico_anual']))
    data['avan_fisico_sem'] = safe_decimal(get_cell(row, col_map['avan_fisico_sem']))

    # Campos codigo.nombre
    for field in ('programa_pptal', 'tipo_prod_proy', 'producto_proyecto',
                  'tipo_act_obra_ac', 'activ_obra_accinv', 'funcion',
                  'division_fn', 'grupo_fn', 'finalidad', 'unidad_medida',
                  'fuente_financ', 'rubro', 'categoria_gasto'):
        raw = safe_str(get_cell(row, col_map[field]))
        cod, nom = parse_codigo_nombre(raw)
        data[f'{field}_cod'] = cod
        data[f'{field}_nom'] = nom
        data[f'{field}_raw'] = raw

    # Filtro restringido (solo en .xlsx, col 101)
    if is_xlsx and 'filtro_restringido' in col_map:
        restringido_val = get_cell(row, col_map['filtro_restringido'])
        data['restringido'] = bool(restringido_val and safe_int(restringido_val))
    else:
        data['restringido'] = False

    # Clasificador de gasto
    if is_xlsx:
        # .xlsx: código y nombre en columnas separadas
        tipo_trans_raw = safe_str(get_cell(row, col_map['tipo_transaccion']))
        tipo_trans_cod, tipo_trans_nom = parse_codigo_nombre(tipo_trans_raw)
        gen_cod = safe_str(get_cell(row, col_map['generica_cod']))
        gen_nom = safe_str(get_cell(row, col_map['generica_nom']))
        subgen_cod = safe_str(get_cell(row, col_map['subgenerica_cod']))
        subgen_nom = safe_str(get_cell(row, col_map['subgenerica_nom']))
        subgen_det_cod = safe_str(get_cell(row, col_map['subgenerica_det_cod']))
        subgen_det_nom = safe_str(get_cell(row, col_map['subgenerica_det_nom']))
        esp_cod = safe_str(get_cell(row, col_map['especifica_cod']))
        esp_nom = safe_str(get_cell(row, col_map['especifica_nom']))
        esp_det_cod = safe_str(get_cell(row, col_map['especifica_det_cod']))
        esp_det_nom = safe_str(get_cell(row, col_map['especifica_det_nom']))
    else:
        # .xls: todo junto "CODIGO.NOMBRE"
        tipo_trans_raw = safe_str(get_cell(row, col_map['tipo_transaccion']))
        tipo_trans_cod, tipo_trans_nom = parse_codigo_nombre(tipo_trans_raw)
        gen_cod, gen_nom = parse_codigo_nombre(safe_str(get_cell(row, col_map['generica'])))
        subgen_cod, subgen_nom = parse_codigo_nombre(safe_str(get_cell(row, col_map['subgenerica'])))
        subgen_det_cod, subgen_det_nom = parse_codigo_nombre(safe_str(get_cell(row, col_map['subgenerica_det'])))
        esp_cod, esp_nom = parse_codigo_nombre(safe_str(get_cell(row, col_map['especifica'])))
        esp_det_cod, esp_det_nom = parse_codigo_nombre(safe_str(get_cell(row, col_map['especifica_det'])))

    data['clasificador_codigo'] = f'{tipo_trans_cod}.{gen_cod}.{subgen_cod}.{subgen_det_cod}.{esp_cod}.{esp_det_cod}'
    data['clasificador_parts'] = {
        'tipo_transaccion': tipo_trans_cod,
        'generica': gen_cod,
        'subgenerica': subgen_cod,
        'subgenerica_det': subgen_det_cod,
        'especifica': esp_cod,
        'especifica_det': esp_det_cod,
        'nombre_tipo_transaccion': tipo_trans_nom[:100],
        'nombre_generica': gen_nom[:200],
        'nombre_subgenerica': subgen_nom[:200],
        'nombre_subgenerica_det': subgen_det_nom[:200],
        'nombre_especifica': esp_nom[:200],
        'nombre_especifica_det': esp_det_nom[:300],
    }

    # Montos
    data['pia'] = safe_decimal(get_cell(row, col_map['mto_pia']))
    data['modificaciones'] = safe_decimal(get_cell(row, col_map['mto_modificaciones']))
    data['pim'] = safe_decimal(get_cell(row, col_map['mto_pim']))
    data['certificado'] = safe_decimal(get_cell(row, col_map['mto_certificado']))
    data['compromiso_anual'] = safe_decimal(get_cell(row, col_map['mto_compro_anual']))

    # Mensuales
    data['mensuales'] = {}
    for mes in range(1, 13):
        data['mensuales'][mes] = {
            'compromiso': safe_decimal(get_cell(row, col_map['mto_at_comp_01'] + (mes - 1))),
            'devengado': safe_decimal(get_cell(row, col_map['mto_devenga_01'] + (mes - 1))),
            'girado': safe_decimal(get_cell(row, col_map['mto_girado_01'] + (mes - 1))),
            'pagado': safe_decimal(get_cell(row, col_map['mto_pagado_01'] + (mes - 1))),
        }

    return data


def _get_or_create_clasificador(data, clasificadores_cache):
    """Obtiene o crea un ClasificadorGasto a partir de los datos parseados."""
    codigo = data['clasificador_codigo']

    if codigo in clasificadores_cache:
        return clasificadores_cache[codigo]

    parts = data['clasificador_parts']
    nombre = parts['nombre_especifica_det'] or parts['nombre_generica'] or codigo

    obj, _ = ClasificadorGasto.objects.get_or_create(
        codigo=codigo[:20],
        defaults={
            'nombre': nombre[:300],
            'tipo_transaccion': parts['tipo_transaccion'],
            'generica': parts['generica'],
            'subgenerica': parts['subgenerica'],
            'subgenerica_det': parts['subgenerica_det'],
            'especifica': parts['especifica'],
            'especifica_det': parts['especifica_det'],
            'nombre_tipo_transaccion': parts['nombre_tipo_transaccion'],
            'nombre_generica': parts['nombre_generica'],
            'nombre_subgenerica': parts['nombre_subgenerica'],
            'nombre_subgenerica_det': parts['nombre_subgenerica_det'],
            'nombre_especifica': parts['nombre_especifica'],
            'nombre_especifica_det': parts['nombre_especifica_det'],
        }
    )
    clasificadores_cache[codigo] = obj
    return obj


def _get_or_create_rubro(data, rubros_cache):
    """Obtiene o crea un Rubro (y su FuenteFinanciamiento padre) a partir de los datos parseados."""
    rubro_cod = data['rubro_cod']

    if rubro_cod in rubros_cache:
        return rubros_cache[rubro_cod]

    # Primero asegurar que la fuente padre existe
    fuente_cod = data['fuente_financ_cod']
    fuente_obj, _ = FuenteFinanciamiento.objects.get_or_create(
        codigo=fuente_cod[:10],
        defaults={
            'nombre': data['fuente_financ_nom'][:200],
            'nombre_corto': fuente_cod,
        }
    )

    # Crear/obtener el rubro
    obj, _ = Rubro.objects.get_or_create(
        codigo=rubro_cod[:10],
        defaults={
            'nombre': data['rubro_nom'][:200],
            'nombre_corto': rubro_cod,
            'fuente': fuente_obj,
        }
    )
    rubros_cache[rubro_cod] = obj
    return obj


def _process_sheetgasto(rows, col_map, is_xlsx, anio_fiscal, importacion, user,
                        sf_mapping=None, rubros_cache=None, clasificadores_cache=None,
                        is_initial=False):
    """
    Procesa SheetGasto con lógica incremental.
    - Carga inicial: metas ya creadas por SF, solo crea ejecuciones. Actualiza avance físico.
    - Daily: solo obtiene metas/clasificadores existentes, crea/actualiza ejecuciones.
    """
    if rubros_cache is None:
        rubros_cache = {r.codigo: r for r in Rubro.objects.all()}
    if clasificadores_cache is None:
        clasificadores_cache = {c.codigo: c for c in ClasificadorGasto.objects.all()}

    stats = {
        'total': len(rows), 'created': 0, 'updated': 0,
        'unchanged': 0, 'errors': 0,
    }
    errors_log = []

    # Cache de metas por sec_func
    if is_initial and sf_mapping:
        # Carga inicial: metas ya creadas por _process_sf_sheet
        metas_cache = {}
        for sec_func, mapping in sf_mapping.items():
            meta_codigo = str(sec_func).zfill(4)
            meta_key = f'{anio_fiscal.id}_{meta_codigo}'
            metas_cache[meta_key] = mapping['meta']
    else:
        # Daily: cargar metas existentes de BD
        metas_cache = {}
        for meta in Meta.objects.filter(anio_fiscal=anio_fiscal):
            meta_key = f'{anio_fiscal.id}_{meta.codigo}'
            metas_cache[meta_key] = meta

    # Cargar existentes en memoria
    existing_ejec = {}
    for ejec in EjecucionPresupuestal.objects.filter(
        anio_fiscal=anio_fiscal
    ).select_related('meta', 'rubro', 'clasificador_gasto'):
        key = build_ejecucion_key(
            ejec.meta.codigo,
            ejec.rubro.codigo,
            ejec.clasificador_gasto.codigo,
        )
        existing_ejec[key] = ejec

    existing_mensuales = {}
    for mens in EjecucionMensual.objects.filter(
        ejecucion__anio_fiscal=anio_fiscal
    ):
        if mens.ejecucion_id not in existing_mensuales:
            existing_mensuales[mens.ejecucion_id] = {}
        existing_mensuales[mens.ejecucion_id][mens.mes] = mens

    # Para actualizar avance físico y cant_meta_anual desde SheetGasto (primera vez que vemos cada meta)
    avance_updated = set()

    # Acumuladores bulk
    ejec_to_create = []
    ejec_to_update = []
    mens_to_create = []
    mens_to_update = []

    for row_num, row in enumerate(rows, start=2):
        try:
            data = _extract_row_data(row, col_map, is_xlsx=is_xlsx)

            sec_func = data['sec_func']
            if sec_func is None:
                continue

            # Meta por sec_func
            meta_codigo = str(sec_func).zfill(4)
            meta_key = f'{anio_fiscal.id}_{meta_codigo}'

            if meta_key not in metas_cache:
                if is_initial:
                    # En carga inicial, la meta debería existir de SF. Si no, skip.
                    stats['errors'] += 1
                    errors_log.append(f'Fila {row_num}: Meta sec_func={sec_func} no encontrada en SF')
                    continue
                else:
                    # En daily, si la meta no existe, skip
                    stats['errors'] += 1
                    errors_log.append(f'Fila {row_num}: Meta sec_func={sec_func} no existe en BD')
                    continue

            meta = metas_cache[meta_key]

            # Actualizar avance físico, cant_meta_anual y tipo_actividad desde SheetGasto (solo primera vez por meta)
            if is_initial and meta_key not in avance_updated:
                avance_updated.add(meta_key)
                # Actualizar cant_meta_anual, unidad_medida y tipo_actividad en la meta
                update_fields = {}
                if data['cant_meta_anual'] > 0 or data['unidad_medida_cod']:
                    update_fields['cantidad_meta_anual'] = data['cant_meta_anual']
                    update_fields['codigo_unidad_medida'] = data['unidad_medida_cod'][:10]
                    update_fields['nombre_unidad_medida'] = data['unidad_medida_nom'][:100]
                # tipo_actividad viene de tipo_act_obra_ac (ACTIVIDAD/OBRA/ACCIÓN DE INVERSIÓN)
                tipo_act = _normalize_tipo_actividad(data.get('tipo_act_obra_ac_nom', ''))
                if tipo_act:
                    update_fields['tipo_actividad'] = tipo_act
                if update_fields:
                    Meta.objects.filter(pk=meta.pk).update(**update_fields)
                # Actualizar AvanceFisico
                AvanceFisico.objects.filter(meta=meta).update(
                    cantidad_meta_semestral=data['cant_meta_sem'],
                    avance_fisico_anual=data['avan_fisico_anual'],
                    avance_fisico_semestral=data['avan_fisico_sem'],
                )

            # Rubro
            if is_initial:
                rubro = _get_or_create_rubro(data, rubros_cache)
            else:
                rubro_cod = data['rubro_cod']
                if rubro_cod in rubros_cache:
                    rubro = rubros_cache[rubro_cod]
                else:
                    # En daily, intentar obtener de BD
                    try:
                        rubro = Rubro.objects.get(codigo=rubro_cod[:10])
                        rubros_cache[rubro_cod] = rubro
                    except Rubro.DoesNotExist:
                        stats['errors'] += 1
                        errors_log.append(f'Fila {row_num}: Rubro {rubro_cod} no existe')
                        continue

            # Clasificador
            if is_initial:
                clasificador = _get_or_create_clasificador(data, clasificadores_cache)
            else:
                clasif_codigo = data['clasificador_codigo']
                if clasif_codigo in clasificadores_cache:
                    clasificador = clasificadores_cache[clasif_codigo]
                else:
                    # En daily, intentar obtener de BD
                    try:
                        clasificador = ClasificadorGasto.objects.get(codigo=clasif_codigo[:20])
                        clasificadores_cache[clasif_codigo] = clasificador
                    except ClasificadorGasto.DoesNotExist:
                        stats['errors'] += 1
                        errors_log.append(f'Fila {row_num}: Clasificador {clasif_codigo} no existe')
                        continue

            # Clave de ejecución
            ejec_key = build_ejecucion_key(meta_codigo, rubro.codigo, clasificador.codigo)
            new_montos = {
                'pia': data['pia'],
                'modificaciones': data['modificaciones'],
                'pim': data['pim'],
                'certificado': data['certificado'],
                'compromiso_anual': data['compromiso_anual'],
            }

            if ejec_key in existing_ejec:
                # EXISTE — verificar si montos cambiaron
                ejec_obj = existing_ejec[ejec_key]
                montos_diff = montos_changed(ejec_obj, new_montos)

                if montos_diff:
                    ejec_obj.pia = data['pia']
                    ejec_obj.modificaciones = data['modificaciones']
                    ejec_obj.pim = data['pim']
                    ejec_obj.certificado = data['certificado']
                    ejec_obj.compromiso_anual = data['compromiso_anual']
                    ejec_obj.codigo_categoria_gasto = data['categoria_gasto_cod'][:10]
                    ejec_obj.nombre_categoria_gasto = data['categoria_gasto_nom'][:100]
                    ejec_obj.restringido = data['restringido']
                    ejec_obj.archivo_origen = importacion
                    ejec_obj.importado_por = user
                    ejec_to_update.append(ejec_obj)
                    stats['updated'] += 1
                else:
                    stats['unchanged'] += 1

                # Mensuales — comparar
                ejec_mensuales = existing_mensuales.get(ejec_obj.id, {})
                for mes in range(1, 13):
                    new_mens_data = data['mensuales'][mes]
                    if mes in ejec_mensuales:
                        if mensual_changed(ejec_mensuales[mes], new_mens_data):
                            obj = ejec_mensuales[mes]
                            obj.compromiso = new_mens_data['compromiso']
                            obj.devengado = new_mens_data['devengado']
                            obj.girado = new_mens_data['girado']
                            obj.pagado = new_mens_data['pagado']
                            mens_to_update.append(obj)
                    else:
                        mens_to_create.append(EjecucionMensual(
                            ejecucion=ejec_obj, mes=mes, **new_mens_data
                        ))
            else:
                # NUEVO — acumular para bulk_create
                ejec_obj = EjecucionPresupuestal(
                    anio_fiscal=anio_fiscal,
                    meta=meta,
                    rubro=rubro,
                    clasificador_gasto=clasificador,
                    codigo_categoria_gasto=data['categoria_gasto_cod'][:10],
                    nombre_categoria_gasto=data['categoria_gasto_nom'][:100],
                    restringido=data['restringido'],
                    archivo_origen=importacion,
                    importado_por=user,
                    **new_montos,
                )
                ejec_obj._pending_mensuales = data['mensuales']
                ejec_to_create.append(ejec_obj)
                stats['created'] += 1

        except Exception as e:
            stats['errors'] += 1
            error_msg = f'Fila {row_num}: {str(e)}'
            errors_log.append(error_msg)
            logger.warning(error_msg)
            if stats['errors'] > 100:
                errors_log.append('Demasiados errores. Importación detenida.')
                break

    # Bulk operations
    if ejec_to_create:
        # Crear ejecuciones una por una para obtener IDs (MySQL no retorna IDs en bulk_create)
        for ejec in ejec_to_create:
            pending = ejec._pending_mensuales if hasattr(ejec, '_pending_mensuales') else None
            if hasattr(ejec, '_pending_mensuales'):
                del ejec._pending_mensuales
            ejec.save()
            if pending:
                for mes, mdata in pending.items():
                    mens_to_create.append(EjecucionMensual(
                        ejecucion_id=ejec.pk, mes=mes, **mdata
                    ))

    if ejec_to_update:
        EjecucionPresupuestal.objects.bulk_update(
            ejec_to_update,
            ['pia', 'modificaciones', 'pim', 'certificado', 'compromiso_anual',
             'codigo_categoria_gasto', 'nombre_categoria_gasto', 'restringido',
             'archivo_origen', 'importado_por'],
            batch_size=500,
        )

    if mens_to_create:
        EjecucionMensual.objects.bulk_create(mens_to_create, batch_size=1000)

    if mens_to_update:
        EjecucionMensual.objects.bulk_update(
            mens_to_update,
            ['compromiso', 'devengado', 'girado', 'pagado'],
            batch_size=1000,
        )

    stats['errors_log'] = errors_log
    return stats


# =============================================================================
# Flujo principal: Carga Inicial (.xlsx con múltiples hojas)
# =============================================================================

def process_initial_load(wb, importacion, user):
    """Procesa el .xlsx completo: RB → CG → Enriquecer CG → SF (+ Metas) → SheetGasto."""
    anio_fiscal = importacion.anio_fiscal
    estadisticas = {}

    # Fase 1: RB (crea FuenteFinanciamiento + Rubro)
    logger.info('Procesando hoja RB...')
    rubros_cache = _process_rb_sheet(wb)
    estadisticas['rubros_procesados'] = len(rubros_cache)

    # Fase 2: CG (crea todos los ClasificadorGasto)
    logger.info('Procesando hoja CG...')
    clasificadores_cache = _process_cg_sheet(wb)
    estadisticas['cg_procesadas'] = len(clasificadores_cache)

    # Fase 3: Enriquecer clasificadores con nombres de SheetGasto
    logger.info('Enriqueciendo clasificadores con nombres de SheetGasto...')
    _enrich_clasificadores_from_sheetgasto(wb, clasificadores_cache)

    # Fase 4: SF (crea UnidadOrganica + Metas)
    logger.info('Procesando hoja SF...')
    sf_mapping = _process_sf_sheet(wb, anio_fiscal)
    estadisticas['sf_procesadas'] = len(sf_mapping)
    estadisticas['metas_creadas'] = len(sf_mapping)

    # Fase 5: SheetGasto (solo ejecuciones, metas ya existen)
    if 'SheetGasto' not in wb.sheetnames:
        raise ValueError('No se encontró la hoja SheetGasto')

    ws = wb['SheetGasto']
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    logger.info(f'Procesando SheetGasto ({len(rows)} filas) con carga inicial...')
    stats = _process_sheetgasto(
        rows, COL_XLSX, is_xlsx=True, anio_fiscal=anio_fiscal,
        importacion=importacion, user=user,
        sf_mapping=sf_mapping, rubros_cache=rubros_cache,
        clasificadores_cache=clasificadores_cache,
        is_initial=True,
    )

    wb.close()

    estadisticas.update({
        'sheetgasto_creadas': stats['created'],
        'sheetgasto_actualizadas': stats['updated'],
        'sheetgasto_sin_cambios': stats['unchanged'],
    })

    return stats, estadisticas


# =============================================================================
# Flujo principal: Importación Diaria (.xls)
# =============================================================================

def process_daily_import(path, importacion, user):
    """
    Procesa el .xls diario: solo SheetGasto con lógica incremental.
    NO modifica metas, clasificadores, rubros ni unidades orgánicas.
    Solo crea/actualiza EjecucionPresupuestal y EjecucionMensual.
    """
    import pandas as pd

    anio_fiscal = importacion.anio_fiscal

    # Leer con pandas (soporte nativo .xls via xlrd)
    try:
        df = pd.read_excel(path, engine='xlrd')
    except Exception:
        df = pd.read_excel(path)

    # Convertir DataFrame a lista de tuplas (compatibilidad con get_cell)
    rows = [tuple(row) for row in df.values]

    logger.info(f'Procesando importación diaria ({len(rows)} filas)...')
    stats = _process_sheetgasto(
        rows, COL_DAILY, is_xlsx=False, anio_fiscal=anio_fiscal,
        importacion=importacion, user=user,
        is_initial=False,
    )

    estadisticas = {
        'sheetgasto_creadas': stats['created'],
        'sheetgasto_actualizadas': stats['updated'],
        'sheetgasto_sin_cambios': stats['unchanged'],
    }

    return stats, estadisticas


# =============================================================================
# Entry point
# =============================================================================

def process_excel_file(importacion_archivo, user=None):
    """
    Punto de entrada principal. Detecta el tipo de archivo y ejecuta el flujo correspondiente.
    - .xlsx con hoja SF → carga inicial (multi-hoja)
    - .xls o .xlsx sin SF → importación diaria (solo SheetGasto)
    """
    importacion = importacion_archivo
    importacion.estado = 'PROCESANDO'
    importacion.save()

    try:
        path = importacion.archivo.path
        is_initial = False

        # Detectar formato
        if path.lower().endswith('.xlsx'):
            wb = load_workbook(path, read_only=True, data_only=True)
            if 'SF' in wb.sheetnames:
                is_initial = True
            else:
                wb.close()

        if is_initial:
            stats, estadisticas = process_initial_load(wb, importacion, user)
        else:
            stats, estadisticas = process_daily_import(path, importacion, user)

        # Actualizar importación
        importacion.total_filas = stats['total']
        importacion.filas_procesadas = stats['created'] + stats['updated'] + stats['unchanged']
        importacion.filas_con_error = stats['errors']
        importacion.filas_creadas = stats['created']
        importacion.filas_actualizadas = stats['updated']
        importacion.filas_sin_cambios = stats['unchanged']
        importacion.log_errores = stats.get('errors_log', [])
        importacion.estadisticas_detalle = estadisticas
        importacion.fecha_fin = timezone.now()

        if stats['errors'] == 0:
            importacion.estado = 'COMPLETADO'
        elif stats['created'] + stats['updated'] > 0:
            importacion.estado = 'PARCIAL'
        else:
            importacion.estado = 'ERROR'

        importacion.save()

        # Generar alertas
        if importacion.filas_procesadas > 0:
            try:
                anio_fiscal = importacion.anio_fiscal
                Alerta.objects.filter(anio_fiscal=anio_fiscal, estado='ACTIVA').delete()
                alertas = generar_alertas(anio_fiscal)
                logger.info(f'Alertas generadas: {len(alertas)}')
            except Exception as e:
                logger.warning(f'Error al generar alertas: {str(e)}')

        logger.info(
            f'Importación completada: {stats["created"]} creadas, '
            f'{stats["updated"]} actualizadas, {stats["unchanged"]} sin cambios, '
            f'{stats["errors"]} errores de {stats["total"]} total'
        )
        return True

    except Exception as e:
        importacion.estado = 'ERROR'
        importacion.log_errores = [f'Error general: {str(e)}']
        importacion.fecha_fin = timezone.now()
        importacion.save()
        logger.error(f'Error en importación: {str(e)}', exc_info=True)
        return False
