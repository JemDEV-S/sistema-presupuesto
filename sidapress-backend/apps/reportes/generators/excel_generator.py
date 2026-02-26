import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone


HEADER_FILL = PatternFill(start_color='1565c0', end_color='1565c0', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
ALT_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
MONEY_FORMAT = '#,##0.00'
PCT_FORMAT = '0.00"%"'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def _style_data(ws, start_row, end_row, num_cols, money_cols=None, pct_cols=None):
    money_cols = money_cols or []
    pct_cols = pct_cols or []
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.font = Font(name='Calibri', size=9)
            if (row - start_row) % 2 == 1:
                cell.fill = ALT_FILL
            if col in money_cols:
                cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal='right')
            elif col in pct_cols:
                cell.number_format = PCT_FORMAT
                cell.alignment = Alignment(horizontal='right')


def generar_reporte_resumen_excel(resumen, anio):
    """Genera Excel con resumen general."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen General'

    # Título
    ws.merge_cells('A1:B1')
    ws['A1'] = f'Resumen de Ejecución Presupuestal - Año Fiscal {anio}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14)
    ws['A2'] = f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(name='Calibri', size=9, color='808080')

    # Headers
    ws['A4'] = 'Concepto'
    ws['B4'] = 'Monto (S/)'
    _style_header(ws, 4, 2)

    data = [
        ('PIA', resumen.get('pia', 0)),
        ('PIM', resumen.get('pim', 0)),
        ('Certificado', resumen.get('certificado', 0)),
        ('Compromiso Anual', resumen.get('compromiso_anual', 0)),
        ('Devengado', resumen.get('devengado', 0)),
        ('Girado', resumen.get('girado', 0)),
        ('Pagado', resumen.get('pagado', 0)),
        ('Saldo PIM', resumen.get('saldo_pim', 0)),
    ]

    for i, (concepto, monto) in enumerate(data, start=5):
        ws.cell(row=i, column=1, value=concepto)
        ws.cell(row=i, column=2, value=float(monto) if monto else 0)

    _style_data(ws, 5, 5 + len(data) - 1, 2, money_cols=[2])

    # Indicadores
    row = 5 + len(data) + 2
    ws.cell(row=row, column=1, value='Indicador')
    ws.cell(row=row, column=2, value='Valor')
    _style_header(ws, row, 2)

    indicators = [
        ('Avance Devengado/PIM', resumen.get('avance_devengado_pct', 0)),
        ('Avance Certificado/PIM', resumen.get('avance_certificado_pct', 0)),
        ('Total Metas', resumen.get('total_metas', 0)),
        ('Total Ejecuciones', resumen.get('total_ejecuciones', 0)),
    ]

    for i, (ind, val) in enumerate(indicators, start=row + 1):
        ws.cell(row=i, column=1, value=ind)
        ws.cell(row=i, column=2, value=float(val) if val else 0)

    _style_data(ws, row + 1, row + len(indicators), 2)

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_reporte_por_unidad_excel(datos_unidades, anio):
    """Genera Excel con ejecución por unidad orgánica."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Por Unidad Orgánica'

    ws.merge_cells('A1:E1')
    ws['A1'] = f'Ejecución por Unidad Orgánica - Año Fiscal {anio}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14)

    headers = ['Unidad Orgánica', 'PIM (S/)', 'Certificado (S/)', 'Devengado (S/)', 'Avance %']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header(ws, 3, 5)

    for i, u in enumerate(datos_unidades, start=4):
        ws.cell(row=i, column=1, value=u.get('unidad_nombre', '-'))
        ws.cell(row=i, column=2, value=float(u.get('total_pim', 0)))
        ws.cell(row=i, column=3, value=float(u.get('total_certificado', 0)))
        ws.cell(row=i, column=4, value=float(u.get('total_devengado', 0)))
        ws.cell(row=i, column=5, value=float(u.get('avance_pct', 0)))

    end_row = 3 + len(datos_unidades)
    _style_data(ws, 4, end_row, 5, money_cols=[2, 3, 4], pct_cols=[5])

    ws.column_dimensions['A'].width = 35
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions['E'].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_reporte_metas_excel(datos_metas, anio):
    """Genera Excel con reporte de metas."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Metas'

    ws.merge_cells('A1:F1')
    ws['A1'] = f'Reporte de Metas Presupuestales - Año Fiscal {anio}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14)

    headers = ['Código', 'Meta', 'Unidad', 'PIM (S/)', 'Devengado (S/)', 'Avance %']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header(ws, 3, 6)

    for i, m in enumerate(datos_metas, start=4):
        ws.cell(row=i, column=1, value=m.get('meta_codigo', '-'))
        ws.cell(row=i, column=2, value=m.get('meta_nombre', '-'))
        ws.cell(row=i, column=3, value=m.get('unidad_nombre', '-'))
        ws.cell(row=i, column=4, value=float(m.get('total_pim', 0)))
        ws.cell(row=i, column=5, value=float(m.get('total_devengado', 0)))
        ws.cell(row=i, column=6, value=float(m.get('avance_pct', 0)))

    end_row = 3 + len(datos_metas)
    _style_data(ws, 4, end_row, 6, money_cols=[4, 5], pct_cols=[6])

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_reporte_tendencia_excel(datos_tendencia, anio):
    """Genera Excel con la tendencia mensual."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tendencia Mensual'

    ws.merge_cells('A1:G1')
    ws['A1'] = f'Ejecución Mensual - Año Fiscal {anio}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14)

    headers = ['Mes', 'Compromiso', 'Devengado', 'Girado',
               'Acum. Compromiso', 'Acum. Devengado', 'Acum. Girado']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header(ws, 3, 7)

    for i, d in enumerate(datos_tendencia, start=4):
        ws.cell(row=i, column=1, value=d.get('mes_nombre', '-'))
        ws.cell(row=i, column=2, value=float(d.get('compromiso', 0)))
        ws.cell(row=i, column=3, value=float(d.get('devengado', 0)))
        ws.cell(row=i, column=4, value=float(d.get('girado', 0)))
        ws.cell(row=i, column=5, value=float(d.get('acum_compromiso', 0)))
        ws.cell(row=i, column=6, value=float(d.get('acum_devengado', 0)))
        ws.cell(row=i, column=7, value=float(d.get('acum_girado', 0)))

    end_row = 3 + len(datos_tendencia)
    _style_data(ws, 4, end_row, 7, money_cols=[2, 3, 4, 5, 6, 7])

    ws.column_dimensions['A'].width = 10
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
